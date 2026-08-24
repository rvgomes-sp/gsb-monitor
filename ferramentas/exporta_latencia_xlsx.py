#!/usr/bin/env python3
"""Exporta as observações de latência (perfil temporal) para Excel — análise externa.

Lê .scratch/latency_observations.jsonl e gera um .xlsx com abas alinhadas à
leitura de 3 níveis (origem -> estabilidade -> atribuição) + outliers.

Uso: python ferramentas/exporta_latencia_xlsx.py [saida.xlsx]
"""
from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCR = Path(__file__).resolve().parents[1] / ".scratch"
BUCKETS = ["D0", "D1", "D2", "D3_PLUS", "ANOMALIA"]


def confianca(n: int) -> str:
    if n < 10:
        return "INDICIO"
    if n < 50:
        return "PRELIMINAR"
    if n < 200:
        return "PADRAO_PROVAVEL"
    return "PADRAO_FORTE"


def carregar() -> list[dict]:
    p = SCR / "latency_observations.jsonl"
    return [json.loads(l) for l in p.open(encoding="utf-8")] if p.exists() else []


HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F3864")


def _cabecalho(ws, cols):
    ws.append(cols)
    for i, _ in enumerate(cols, 1):
        c = ws.cell(row=1, column=i)
        c.font = HDR
        c.fill = FILL
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autolarg(ws, maxw=60):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(maxw, w + 2)


def gerar(obs: list[dict], saida: Path):
    wb = openpyxl.Workbook()

    # 1) Observações (raw)
    ws = wb.active
    ws.title = "Observacoes"
    cols = ["dia_coleta", "source_sender_raw", "source_host", "source_type", "org_cnpj",
            "org_name", "uf", "modalidade_id", "modalidade", "data_resultado", "data_inclusao",
            "delta_days", "delta_bucket", "inclusion_hour", "result_key"]
    _cabecalho(ws, cols)
    for o in sorted(obs, key=lambda x: (x["source_sender_raw"], x["dia_coleta"])):
        ws.append([o.get(c) for c in cols])
    _autolarg(ws)

    # 2) Por origem (nível 1+2)
    ws = wb.create_sheet("Por_origem")
    _cabecalho(ws, ["source_sender_raw", "source_type", "n", "dias_distintos", "multi_dia",
                    "D0", "D1", "D2", "D3_PLUS", "ANOMALIA", "pct_D0", "confianca", "hora_top"])
    por = defaultdict(lambda: Counter())
    dias = defaultdict(set)
    hora = defaultdict(Counter)
    stype = {}
    for o in obs:
        s = o["source_sender_raw"]
        por[s][o["delta_bucket"]] += 1
        por[s]["n"] += 1
        dias[s].add(o["dia_coleta"])
        hora[s][o["inclusion_hour"]] += 1
        stype[s] = o["source_type"]
    for s in sorted(por, key=lambda k: -por[k]["n"]):
        c = por[s]
        n = c["n"]
        ws.append([s, stype[s], n, len(dias[s]), len(dias[s]) > 1,
                   c["D0"], c["D1"], c["D2"], c["D3_PLUS"], c["ANOMALIA"],
                   round(100 * c["D0"] / n, 1) if n else 0, confianca(n),
                   ", ".join(f"{h}h({q})" for h, q in hora[s].most_common(4))])
    _autolarg(ws)

    # 3) Origem x Órgão / x Dia / x Modalidade (nível 3 - atribuição)
    for titulo, chave in (("Origem_x_Orgao", "org_name"),
                          ("Origem_x_Dia", "dia_coleta"),
                          ("Origem_x_Modalidade", "modalidade")):
        ws = wb.create_sheet(titulo)
        _cabecalho(ws, ["source_sender_raw", chave, "n", "D0", "D1", "D2", "D3_PLUS", "ANOMALIA", "pct_D0"])
        cross = defaultdict(lambda: Counter())
        for o in obs:
            k = (o["source_sender_raw"], o.get(chave))
            cross[k][o["delta_bucket"]] += 1
            cross[k]["n"] += 1
        for (s, k), c in sorted(cross.items(), key=lambda kv: (kv[0][0], -kv[1]["n"])):
            n = c["n"]
            ws.append([s, k, n, c["D0"], c["D1"], c["D2"], c["D3_PLUS"], c["ANOMALIA"],
                       round(100 * c["D0"] / n, 1) if n else 0])
        _autolarg(ws)

    # 4) Outliers (D2 / D3_PLUS / ANOMALIA) — investigação individual
    ws = wb.create_sheet("Outliers_D2_D3_ANOM")
    cols = ["delta_bucket", "delta_days", "dia_coleta", "source_sender_raw", "source_host",
            "source_type", "org_name", "uf", "modalidade", "data_resultado", "data_inclusao",
            "inclusion_hour", "result_key"]
    _cabecalho(ws, cols)
    out = [o for o in obs if o["delta_bucket"] in ("D2", "D3_PLUS", "ANOMALIA")]
    for o in sorted(out, key=lambda x: -x["delta_days"]):
        ws.append([o.get(c) for c in cols])
    _autolarg(ws)

    # 5) Resumo
    ws = wb.create_sheet("Resumo", 0)
    ws.append(["Perfil Temporal EVT-007 — Latência de integração por origem"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["total observações", len(obs)])
    ws.append(["dias", ", ".join(sorted(set(o["dia_coleta"] for o in obs)))])
    ws.append(["origens distintas", len(set(o["source_sender_raw"] for o in obs))])
    tot = Counter(o["delta_bucket"] for o in obs)
    ws.append([])
    ws.append(["bucket", "n", "%"])
    for b in BUCKETS:
        ws.append([b, tot[b], round(100 * tot[b] / len(obs), 1) if obs else 0])
    ws.append([])
    ws.append(["Régua de confiança: <10 INDICIO | 10-49 PRELIMINAR | 50-199 PADRAO_PROVAVEL | >=200 PADRAO_FORTE (exigir multi_dia)"])
    ws.append(["source_type é HEURÍSTICO (usuarioNome + host) — não é atribuição definitiva de plataforma."])
    _autolarg(ws, 90)

    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    print(f"Excel gerado: {saida} | {len(obs)} obs | abas: {wb.sheetnames}")


def main():
    saida = Path(sys.argv[1]) if len(sys.argv) > 1 else (SCR / "GSB_Perfil_Temporal_EVT007.xlsx")
    obs = carregar()
    if not obs:
        print("Sem observações em .scratch/latency_observations.jsonl")
        return
    gerar(obs, saida)


if __name__ == "__main__":
    main()
