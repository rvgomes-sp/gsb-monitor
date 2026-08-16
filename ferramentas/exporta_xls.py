"""
Exporta TODAS as colunas de gsb.evt007_results para XLS.
Uso: python exporta_xls.py
Gera: GSB_Base_Completa_AAAAMMDD.xlsx na pasta atual.
"""
import os, psycopg
from datetime import date
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Instalando openpyxl..."); os.system("pip install openpyxl --quiet"); 
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

DB = os.environ.get("DATABASE_URL","")
if not DB:
    print("ERRO: defina $env:DATABASE_URL antes de rodar."); raise SystemExit(1)

with psycopg.connect(DB) as conn:
    with conn.cursor() as cur:
        # todas as colunas, todas as linhas da base
        cur.execute("SELECT * FROM gsb.evt007_results ORDER BY homologated_total_value DESC NULLS LAST")
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()

print(f"Exportando {len(rows)} linhas x {len(cols)} colunas...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Base Completa"

# cabeçalho
for j, c in enumerate(cols, 1):
    cell = ws.cell(row=1, column=j, value=c)
    cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill("solid", fgColor="1F3864")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# dados
for i, row in enumerate(rows, 2):
    for j, val in enumerate(row, 1):
        # converter tipos que o Excel não aceita direto
        if isinstance(val, (dict, list)):
            val = str(val)
        elif hasattr(val, "isoformat"):
            val = val.isoformat()
        ws.cell(row=i, column=j, value=val)

# largura básica
for j, c in enumerate(cols, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = min(max(len(c)+2, 12), 40)

ws.freeze_panes = "A2"

fname = f"GSB_Base_Completa_{date.today().strftime('%Y%m%d')}.xlsx"
wb.save(fname)
print(f"OK: {fname} ({len(rows)} linhas, {len(cols)} colunas)")
