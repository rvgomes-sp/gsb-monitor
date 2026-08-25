#!/usr/bin/env python3
"""Biblioteca de OBJETOS rotulada — semente do 'catálogo melhorado' (dimensão garantia).

Para cada contratação >= piso do dia, gera:
  - objeto COMPLETO (cru)
  - minha inferência: família + descrição curta + HIPÓTESE de garantia + motivo
  - colunas VAZIAS para o Rodrigo preencher: CORRETO, GARANTIA (SIM/NAO), NOTAS

Saída: planilha .xlsx (editável) ordenada por família e valor. A verdade que o
Rodrigo preencher vira o golden set desta biblioteca (quais objetos/órgãos pedem
garantia, onde, por quê).
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "monitor"))
from inferencia import reduzir_objeto  # noqa: E402

PISO = 10_000_000

# Famílias — ordem importa (primeiro match vence). (regex, família, hipótese_garantia, motivo)
# hipótese: SIM_FORTE | PROVAVEL | DEPENDE | IMPROVAVEL
REGRAS: list[tuple[str, str, str, str]] = [
    (r"concess[ãa]o|\bppp\b|parceria\s+p[úu]blico", "CONCESSAO_PPP", "DEPENDE",
     "concessão/PPP: garantia própria de proposta/execução do contrato de concessão"),
    (r"pavimenta|recapea|asf[áa]lt|cbuq|\btsd\b|terraplen|paralelep[íi]pedo|intertravad|micro\s*revest",
     "OBRA_PAVIMENTACAO", "SIM_FORTE", "obra viária: garantia de execução típica"),
    (r"rodovi|\bbr-?\d|\bma-?\d|\bmt-?\d|\bms-?\d|\bgo-?\d|\bpb-?\d|estrada|duplica[çc]|restaura[çc][ãa]o\s+rodovi|malha\s+rodovi",
     "OBRA_RODOVIA", "SIM_FORTE", "obra rodoviária: garantia de execução típica"),
    (r"ponte|viaduto|passarela|\boae\b|obra\s+de\s+arte", "OBRA_ARTE_ESPECIAL", "SIM_FORTE",
     "arte especial: garantia de execução típica"),
    (r"esgot|\bete\b|\beee\b|adutora|reservat[óo]ri|barragem|drenagem|a[çc]ude|abastecimento\s+de\s+[áa]gua|saneament|\bsaa\b|\bses\b|esta[çc][ãa]o\s+de\s+tratam",
     "OBRA_SANEAMENTO", "SIM_FORTE", "obra de saneamento/hídrica: garantia de execução típica"),
    (r"constru[çc][ãa]o|edifica[çc]|\bupa\b|hospital|escola|creche|\bcei\b|gin[áa]sio|quadra|pr[ée]dio|sede|penitenci|reforma|revitaliza|requalifica|amplia[çc]|adequa[çc]|infraestrutura\s+urban",
     "OBRA_EDIFICACAO", "SIM_FORTE", "edificação/reforma predial: garantia de execução típica"),
    (r"linha\s+de\s+transmiss|subesta[çc]|rede\s+de\s+distribui|fotovoltaic|solar|il?umina[çc][ãa]o\s+p[úu]blica",
     "OBRA_ELETRICA", "SIM_FORTE", "infraestrutura elétrica: geralmente obra/engenharia"),
    (r"servi[çc]os?\s+(comuns?\s+)?de\s+engenharia|engenharia\s+.*manuten[çc]|manuten[çc][ãa]o\s+predial|manuten[çc][ãa]o\s+rodovi|conserva[çc][ãa]o.*(via|logradouro|rodovi)",
     "SERVICO_ENGENHARIA", "PROVAVEL", "serviço de engenharia/manutenção: garantia frequente, depende da escala"),
    (r"limpeza|vigil[âa]ncia|seguran[çc]a|portaria|serventia|conserva[çc][ãa]o\s+e\s+limp|terceiriza|m[ãa]o\s+de\s+obra|brigad|recep[çc]|dedica[çc][ãa]o\s+exclusiv|serventes?",
     "SERVICO_CONTINUO_MAO_OBRA", "PROVAVEL", "serviço contínuo c/ mão de obra: garantia + conta vinculada frequentes"),
    (r"medicament|f[áa]rmac|\bopme\b|[óo]rtese|pr[óo]tese|insumo\s+laborator|material\s+m[ée]dic|material\s+hospitalar|dieta|seringa|odontol[óo]gic|enfermagem|hospitalar\s+diverso|kit\s+para\s+ajuda",
     "MEDICAMENTO_SAUDE", "IMPROVAVEL", "aquisição de saúde (SRP): garantia de execução incomum"),
    (r"alimenta[çc]|g[êe]nero|merenda|nutri[çc]|carne|latic[íi]nio|frango|dieta\s+enteral|produtos\s+n[ãa]o\s+pereciv",
     "ALIMENTACAO", "IMPROVAVEL", "aquisição de alimentos (SRP): garantia incomum"),
    (r"combust[íi]vel|[óo]leo\s+diesel|arla|gasolina", "COMBUSTIVEL", "IMPROVAVEL",
     "fornecimento de combustível: garantia incomum"),
    (r"ve[íi]cul|caminh[ãa]o|caminhonete|motociclet|m[áa]quina|trator|[ôo]nibus|pneu|c[âa]mara\s+de\s+ar",
     "VEICULO_MAQUINA", "IMPROVAVEL", "aquisição de veículos/máquinas: garantia de execução incomum"),
    (r"tom[óo]graf|mobili[áa]ri|ar\s+condicionad|equipament|aparelho|notebook|microcomputad|computador|servidor|tel[ei]ssaude|telessa[úu]de",
     "EQUIPAMENTO_MOBILIARIO", "IMPROVAVEL", "aquisição de equipamento/mobiliário: garantia incomum"),
    (r"software|licenciament|solu[çc][ãa]o\s+integrad|tecnologia\s+da\s+informa|seguran[çc]a\s+da\s+informa|link\s+de\s+telecom|desenvolvimento.*software|sistema\b",
     "TI_SOFTWARE", "DEPENDE", "serviço/solução de TI: garantia varia"),
    (r"tubo|conex[ãa]o|\bpvc\b|\bpead\b|a[çc]os?\s+e\s+ferros?|vergalh[ãa]o|cimento|material.*constru[çc][ãa]o|uniforme|vestu[áa]rio|\bepi\b|descart[áa]vel",
     "MATERIAL_INSUMO", "IMPROVAVEL", "aquisição de material/insumo: garantia de execução incomum"),
    (r"loca[çc][ãa]o", "LOCACAO", "DEPENDE", "locação: garantia varia conforme escopo"),
    (r"publicidade|propaganda|marketing|consultoria|regulariza[çc][ãa]o\s+fundi|estudos?\s+socioamb|ensino\s+superior|\bies\b|transporte\s+escolar|gr[áa]fic|serventia|abastec|gerenciament",
     "SERVICO_DIVERSO", "DEPENDE", "serviço diverso: garantia varia"),
]


_CTRL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _limpo(s):
    return _CTRL.sub(" ", s) if isinstance(s, str) else s


def classificar(objeto: str) -> tuple[str, str, str]:
    low = re.sub(r"\s+", " ", (objeto or "")).lower()
    for pat, fam, gar, mot in REGRAS:
        if re.search(pat, low):
            return fam, gar, mot
    return "OUTRO", "DEPENDE", "não classificado pela regra — revisar"


def main() -> int:
    arq = sys.argv[1] if len(sys.argv) > 1 else "saidas/raw_20260821.json"
    d = json.loads(Path(arq).read_text(encoding="utf-8"))
    linhas = [r for r in d["linhas"] if (float(r.get("valorTotalEstimado") or 0) >= PISO)]

    itens = []
    for r in linhas:
        obj = r.get("objetoCompra") or ""
        fam, gar, mot = classificar(obj)
        itens.append({
            "controle": r.get("numeroControlePNCP"),
            "uf": (r.get("unidadeOrgao") or {}).get("ufSigla"),
            "orgao": (r.get("orgaoEntidade") or {}).get("razaoSocial"),
            "mod": r.get("modalidadeId"),
            "valor_mm": round(float(r.get("valorTotalEstimado") or 0) / 1e6, 2),
            "situacao": r.get("situacaoCompraNome"),
            "objeto": re.sub(r"\s+", " ", obj).strip(),
            "descricao_curta": reduzir_objeto(obj, 80),
            "familia": fam, "garantia_hip": gar, "motivo": mot,
        })
    # ordena por família e valor desc
    itens.sort(key=lambda x: (x["familia"], -x["valor_mm"]))

    # ---- planilha ----
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "biblioteca_objetos"
    cabecalho = ["#", "controle_pncp", "UF", "órgão", "mod", "valor R$MM", "situação",
                 "OBJETO COMPLETO", "família (inferida)", "descrição curta",
                 "garantia? (hipótese)", "motivo da hipótese",
                 ">> CORRETO (você)", ">> GARANTIA SIM/NÃO (você)", ">> NOTAS (você)"]
    ws.append(cabecalho)
    hd = Font(bold=True, color="FFFFFF")
    fill_meu = PatternFill("solid", fgColor="305496")
    fill_dele = PatternFill("solid", fgColor="C55A11")
    for c, nome in enumerate(cabecalho, 1):
        cell = ws.cell(1, c)
        cell.font = hd
        cell.fill = fill_dele if nome.startswith(">>") else fill_meu
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, it in enumerate(itens, 1):
        ws.append([i, it["controle"], it["uf"], _limpo(it["orgao"]), it["mod"], it["valor_mm"],
                   it["situacao"], _limpo(it["objeto"]), it["familia"], _limpo(it["descricao_curta"]),
                   it["garantia_hip"], it["motivo"], "", "", ""])
    larguras = [4, 26, 5, 34, 5, 11, 16, 70, 24, 34, 18, 40, 22, 18, 30]
    for c, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    out = Path("saidas/biblioteca_objetos_20260821.xlsx")
    wb.save(out)
    print(f"planilha: {out} | {len(itens)} objetos")

    # ---- resumo por família (para o chat) ----
    from collections import Counter
    fam_ct = Counter(it["familia"] for it in itens)
    print("\n=== FAMÍLIAS (contagem | hipótese de garantia) ===")
    ghip = {it["familia"]: it["garantia_hip"] for it in itens}
    for fam, ct in sorted(fam_ct.items(), key=lambda x: -x[1]):
        print(f"  {ct:>3}  {fam:28} garantia? {ghip[fam]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
